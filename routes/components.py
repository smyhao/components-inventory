from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any

import chardet
from flask import Blueprint, Response, request
from openpyxl import Workbook, load_workbook

from config import UPLOAD_FOLDER
from models import (
    InventoryError,
    clean_int,
    clean_optional_int,
    clean_text,
    normalize_footprint,
    normalize_value,
)
from routes._utils import excel_response, success

components_bp = Blueprint("components", __name__)


@components_bp.get("/api/components")
def api_components():
    from app import repo

    return success(repo.list_components(request.args.to_dict()))


@components_bp.post("/api/components")
def api_create_component():
    from app import repo

    return success(repo.create_component(request.get_json(silent=True) or {}))


@components_bp.get("/api/components/<int:component_id>")
def api_component_detail(component_id: int):
    from app import repo

    return success(repo.get_component(component_id))


@components_bp.put("/api/components/<int:component_id>")
def api_update_component(component_id: int):
    from app import repo

    return success(repo.update_component(component_id, request.get_json(silent=True) or {}))


@components_bp.delete("/api/components/<int:component_id>")
def api_delete_component(component_id: int):
    from app import repo

    repo.delete_component(component_id)
    return success({"id": component_id})


@components_bp.get("/api/components/<int:component_id>/qr.svg")
def api_component_qr_svg(component_id: int):
    from app import repo

    repo.get_component(component_id)
    return Response(
        _component_qr_svg(component_id),
        mimetype="image/svg+xml",
        headers={"Cache-Control": "no-store"},
    )


@components_bp.post("/api/components/import")
def api_import_components():
    file = request.files.get("file")
    if not file:
        raise InventoryError("missing upload file")
    result = _parse_component_import(file)
    return success(result)


@components_bp.get("/api/components/export")
def api_export_components():
    from app import file_logger

    data = _export_components_xlsx()
    file_logger.write_backend("COMPONENT", "export components excel")
    return excel_response(data, "components_export.xlsx")


# --- Helper functions ---


def _component_link(component_id: int) -> str:
    from config import SERVER_URL
    from flask import request as req

    return f"{(req.url_root or SERVER_URL).rstrip('/')}/?component={component_id}"


def _component_qr_svg(component_id: int) -> bytes:
    try:
        import qrcode
        import qrcode.image.svg
    except ImportError as exc:
        raise InventoryError("missing dependency qrcode, please run pip install -r requirements.txt") from exc

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=2,
        image_factory=qrcode.image.svg.SvgPathImage,
    )
    qr.add_data(_component_link(component_id))
    qr.make(fit=True)
    image = qr.make_image()
    buffer = io.BytesIO()
    image.save(buffer)
    return buffer.getvalue()


def _split_tags(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [clean_text(item) for item in value if clean_text(item)]
    return [item.strip() for item in re.split("[,，]", str(value)) if item.strip()]


def _normalize_header(value: Any) -> str:
    return clean_text(value).strip().lower().replace(" ", "").replace("_", "")


COMPONENT_HEADER_ALIASES = {
    "name": {"名称", "元器件名称", "name"},
    "category": {"分类", "category"},
    "model": {"型号", "model"},
    "package": {"封装", "package", "footprint"},
    "nominal_value": {"标称值", "值", "value", "comment"},
    "voltage_rating": {"耐压", "voltage", "voltage_rating"},
    "current_rating": {"耐流", "电流", "current", "current_rating"},
    "power_rating": {"功率", "power", "power_rating"},
    "tolerance": {"精度", "tolerance"},
    "material_type": {"材质/类型", "材质", "类型", "material", "material_type"},
    "manufacturer": {"厂商", "manufacturer"},
    "quantity": {"数量", "qty", "quantity"},
    "box_name": {"盒子名称", "收纳盒", "box", "box_name"},
    "row": {"行", "row"},
    "col": {"列", "col", "column"},
    "description": {"描述", "备注", "description"},
    "tags": {"标签", "tags"},
}


def _resolve_headers(header_row: list[Any], aliases: dict[str, set[str]]) -> dict[str, int]:
    normalized = {index: _normalize_header(value) for index, value in enumerate(header_row)}
    mapping: dict[str, int] = {}
    for field, names in aliases.items():
        normalized_aliases = {_normalize_header(name) for name in names}
        for index, value in normalized.items():
            if value in normalized_aliases:
                mapping[field] = index
                break
    return mapping


def _workbook_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _decode_csv_bytes(data: bytes) -> str:
    detected = chardet.detect(data)
    candidates = [detected.get("encoding"), "utf-8-sig", "utf-8", "utf-16", "gbk"]
    for encoding in candidates:
        if not encoding:
            continue
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def _parse_csv_rows(data: bytes) -> list[list[Any]]:
    text = _decode_csv_bytes(data)
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel_tab if "\t" in sample else csv.excel
    return list(csv.reader(io.StringIO(text), dialect))


def _parse_xlsx_rows(data: bytes) -> list[tuple[Any, ...]]:
    workbook = load_workbook(filename=io.BytesIO(data), data_only=True)
    sheet = workbook.active
    return list(sheet.iter_rows(values_only=True))


def _component_import_rows(file_storage) -> tuple[list[Any], str]:
    filename = file_storage.filename or ""
    ext = Path(filename).suffix.lower()
    data = file_storage.read()
    if ext == ".csv" or ext == ".txt":
        return _parse_csv_rows(data), "csv"
    if ext == ".xlsx":
        return _parse_xlsx_rows(data), "xlsx"
    raise InventoryError("only csv and xlsx files are supported")


def _parse_component_import(file_storage) -> dict[str, Any]:
    from app import file_logger, repo

    rows, file_type = _component_import_rows(file_storage)
    if not rows:
        return {"success": 0, "failed": 0, "errors": [], "file_type": file_type}

    header_map = _resolve_headers(list(rows[0]), COMPONENT_HEADER_ALIASES)
    if "name" not in header_map:
        raise InventoryError("missing required header: name")

    def cell(row: Any, field: str, default: Any = "") -> Any:
        index = header_map.get(field)
        if index is None:
            return default
        return row[index] if index < len(row) else default

    errors: list[dict[str, Any]] = []
    success_count = 0
    failed_count = 0

    for row_index, row in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in row):
            continue
        try:
            name = clean_text(cell(row, "name"))
            if not name:
                raise InventoryError("missing component name")

            category_name = clean_text(cell(row, "category"))
            category_id = repo.ensure_category_by_name(category_name) if category_name else None

            box = None
            compartment = None
            box_name = clean_text(cell(row, "box_name"))
            if box_name:
                box = repo.find_box_by_name(box_name)
                if not box:
                    raise InventoryError(f"box not found: {box_name}")
                grid_row = clean_optional_int(cell(row, "row"))
                grid_col = clean_optional_int(cell(row, "col"))
                if grid_row is not None and grid_col is not None:
                    compartment = repo.find_compartment_by_coordinates(box["id"], grid_row, grid_col)
                    if not compartment:
                        raise InventoryError(f"compartment not found: {box_name} {grid_row}-{grid_col}")

            payload = {
                "name": name,
                "category_id": category_id,
                "model": clean_text(cell(row, "model")),
                "package": clean_text(cell(row, "package")),
                "nominal_value": clean_text(cell(row, "nominal_value")),
                "voltage_rating": clean_text(cell(row, "voltage_rating")),
                "current_rating": clean_text(cell(row, "current_rating")),
                "power_rating": clean_text(cell(row, "power_rating")),
                "tolerance": clean_text(cell(row, "tolerance")),
                "material_type": clean_text(cell(row, "material_type")),
                "manufacturer": clean_text(cell(row, "manufacturer")),
                "quantity": clean_int(cell(row, "quantity"), 0),
                "description": clean_text(cell(row, "description")),
                "tags": _split_tags(cell(row, "tags")) if "tags" in header_map else [],
                "box_id": box["id"] if box else None,
                "compartment_id": compartment["id"] if compartment else None,
                "cell_row": compartment["id"] if compartment else None,
                "cell_col": compartment["col"] if compartment else None,
            }
            repo.create_component(payload, source="import")
            success_count += 1
        except Exception as exc:
            failed_count += 1
            errors.append({"row": row_index, "reason": str(exc)})

    file_logger.write_backend(
        "COMPONENT",
        f"component {file_type} import finished: success={success_count}, failed={failed_count}",
    )
    return {"success": success_count, "failed": failed_count, "errors": errors, "file_type": file_type}


def _export_components_xlsx() -> bytes:
    from app import repo

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "components"
    sheet.append(
        [
            "名称",
            "分类",
            "型号",
            "封装",
            "标称值",
            "耐压",
            "耐流",
            "功率",
            "精度",
            "材质/类型",
            "厂商",
            "数量",
            "收纳盒",
            "行",
            "列",
            "描述",
            "标签",
        ]
    )

    page = 1
    while True:
        result = repo.list_components({"page": page, "page_size": 500})
        for item in result["items"]:
            sheet.append(
                [
                    item.get("name"),
                    item.get("category_name"),
                    item.get("model"),
                    item.get("package"),
                    item.get("nominal_value"),
                    item.get("voltage_rating"),
                    item.get("current_rating"),
                    item.get("power_rating"),
                    item.get("tolerance"),
                    item.get("material_type"),
                    item.get("manufacturer"),
                    item.get("quantity"),
                    item.get("box_name"),
                    item.get("grid_row"),
                    item.get("grid_col"),
                    item.get("description"),
                    ", ".join(item.get("tags") or []),
                ]
            )
        if len(result["items"]) < result["page_size"]:
            break
        page += 1
    return _workbook_bytes(workbook)
