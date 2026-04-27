from __future__ import annotations

import csv
import hmac
from html import escape
import io
import re
import uuid
from pathlib import Path
from typing import Any

import chardet
from flask import Flask, Response, jsonify, request, send_file, send_from_directory
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageOps
from werkzeug.exceptions import HTTPException
from werkzeug.utils import secure_filename

from config import (
    ALLOWED_DOCUMENT_EXTENSIONS,
    ALLOWED_IMAGE_EXTENSIONS,
    API_TOKEN,
    API_TOKEN_REQUIRE_ALL,
    DATABASE_PATH,
    DOCUMENT_FOLDER,
    FRONTEND_LOG_FILE,
    HOST,
    LOG_FILE,
    MAX_CONTENT_LENGTH,
    PORT,
    SERVER_URL,
    STATIC_DIR,
    UPLOAD_FOLDER,
    ensure_runtime_directories,
)
from init_db import init_database
from logger import FileLogger
from models import (
    InventoryError,
    InventoryRepository,
    clean_int,
    clean_optional_int,
    clean_text,
    normalize_footprint,
    normalize_value,
)


ensure_runtime_directories()
init_database(DATABASE_PATH)

app = Flask(__name__, static_folder=str(STATIC_DIR), static_url_path="/static")
app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_LENGTH

file_logger = FileLogger(LOG_FILE, FRONTEND_LOG_FILE)
repo = InventoryRepository(DATABASE_PATH, file_logger, UPLOAD_FOLDER)
file_logger.write_backend("SYSTEM", "backend initialized")


def success(data: Any = None, message: str = "ok", status: int = 200):
    return jsonify({"code": 0, "data": data, "message": message}), status


def failure(message: str, status: int = 400):
    return jsonify({"code": 1, "data": None, "message": message}), status


def excel_response(data: bytes, filename: str):
    return send_file(
        io.BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def current_server_url() -> str:
    return (request.url_root or SERVER_URL).rstrip("/")


def bearer_token() -> str:
    value = request.headers.get("Authorization", "")
    prefix = "Bearer "
    if value.startswith(prefix):
        return value[len(prefix) :].strip()
    return ""


def api_token_identity(token: str) -> dict[str, Any] | None:
    if API_TOKEN and hmac.compare_digest(token, API_TOKEN):
        return {"id": None, "name": "env:API_TOKEN"}
    if token:
        return repo.validate_api_token(token)
    return None


@app.before_request
def require_api_token_for_automation():
    if not API_TOKEN or not request.path.startswith("/api"):
        if not request.path.startswith("/api"):
            return None
        source = request.headers.get("X-Inventory-Source", "")
        should_require_generated = source == "inventory-cli" and repo.has_api_tokens()
        if not should_require_generated:
            return None
        identity = api_token_identity(bearer_token())
        if not identity:
            return failure("authentication failed", 401)
        request.inventory_token = identity
        return None
    source = request.headers.get("X-Inventory-Source", "")
    should_require = API_TOKEN_REQUIRE_ALL or source == "inventory-cli"
    if should_require:
        identity = api_token_identity(bearer_token())
        if not identity:
            return failure("authentication failed", 401)
        request.inventory_token = identity
    return None


@app.after_request
def log_automation_write(response):
    identity = getattr(request, "inventory_token", None)
    if (
        identity
        and request.path.startswith("/api")
        and request.method in {"POST", "PUT", "DELETE"}
        and response.status_code < 500
    ):
        actor = request.headers.get("X-Inventory-Actor") or identity.get("name") or "automation"
        file_logger.write_backend(
            "SYSTEM",
            f"api write: token={identity.get('name')}, actor={actor}, method={request.method}, path={request.path}, status={response.status_code}",
        )
    return response


def component_link(component_id: int) -> str:
    return f"{current_server_url()}/?component={component_id}"


def component_qr_svg(component_id: int) -> bytes:
    try:
        import qrcode
        import qrcode.image.svg
    except ImportError as exc:
        raise InventoryError("missing dependency qrcode, please run pip install -r requirements.txt") from exc

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=2,
        image_factory=qrcode.image.svg.SvgPathImage,
    )
    qr.add_data(component_link(component_id))
    qr.make(fit=True)
    image = qr.make_image()
    buffer = io.BytesIO()
    image.save(buffer)
    return buffer.getvalue()


def parse_component_ids(value: str) -> list[int]:
    ids: list[int] = []
    seen: set[int] = set()
    for part in re.split(r"[,，\s]+", value or ""):
        component_id = clean_int(part, 0)
        if component_id > 0 and component_id not in seen:
            ids.append(component_id)
            seen.add(component_id)
    return ids[:500]


def qr_label_print_page(components: list[dict[str, Any]]) -> str:
    label_count = len(components)
    cards = []
    for item in components:
        component_id = item["id"]
        spec = " / ".join(
            str(value)
            for value in [item.get("model"), item.get("package"), item.get("nominal_value")]
            if value
        )
        location = " / ".join(
            str(value)
            for value in [item.get("box_name"), item.get("cell_label")]
            if value
        )
        quantity = item.get("quantity") if item.get("quantity") is not None else 0
        cards.append(
            f"""
            <article class="label">
                <img src="/api/components/{component_id}/qr.svg" alt="QR">
                <div class="label-text">
                    <strong>{escape(str(item.get("name") or ""))}</strong>
                    <span>{escape(spec or "无规格信息")}</span>
                    <small>{escape(location or "未入盒")} · 库存 {escape(str(quantity))}</small>
                </div>
            </article>
            """
        )
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>元器件二维码标签</title>
    <style>
        * {{ box-sizing: border-box; }}
        body {{ margin: 0; font-family: Arial, "Microsoft YaHei", sans-serif; color: #1f2933; background: #f7f5ef; }}
        header {{ position: sticky; top: 0; z-index: 1; display: flex; align-items: center; justify-content: space-between; gap: 16px; padding: 14px 18px; background: rgba(255,255,255,.94); border-bottom: 1px solid #ddd6c8; }}
        h1 {{ margin: 0; font-size: 18px; }}
        p {{ margin: 4px 0 0; color: #697386; font-size: 13px; }}
        button {{ border: 0; border-radius: 8px; padding: 10px 16px; background: #2563eb; color: white; font-weight: 700; cursor: pointer; }}
        main {{ padding: 18px; }}
        .sheet {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(210px, 1fr)); gap: 10px; }}
        .label {{ min-height: 96px; display: grid; grid-template-columns: 78px minmax(0, 1fr); gap: 10px; align-items: center; padding: 8px; background: white; border: 1px dashed #a8b0bd; border-radius: 8px; break-inside: avoid; }}
        .label img {{ width: 78px; height: 78px; display: block; }}
        .label-text {{ min-width: 0; display: flex; flex-direction: column; gap: 4px; }}
        .label-text strong {{ font-size: 14px; line-height: 1.2; overflow-wrap: anywhere; }}
        .label-text span, .label-text small {{ font-size: 11px; color: #52606d; overflow-wrap: anywhere; }}
        .empty {{ padding: 40px; text-align: center; color: #697386; background: white; border-radius: 8px; }}
        @media print {{
            body {{ background: white; }}
            header {{ display: none; }}
            main {{ padding: 0; }}
            .sheet {{ grid-template-columns: repeat(3, 1fr); gap: 4mm; }}
            .label {{ border-color: #777; border-radius: 2mm; page-break-inside: avoid; }}
        }}
    </style>
</head>
<body>
    <header>
        <div>
            <h1>元器件二维码标签</h1>
            <p>共 {label_count} 个标签。扫码后会打开对应元器件详情，可继续修改参数或库存。</p>
        </div>
        <button onclick="window.print()">打印</button>
    </header>
    <main>
        {"<section class='sheet'>" + "".join(cards) + "</section>" if cards else "<div class='empty'>没有可打印的元器件</div>"}
    </main>
</body>
</html>"""


def split_tags(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [clean_text(item) for item in value if clean_text(item)]
    return [item.strip() for item in re.split("[,\uFF0C]", str(value)) if item.strip()]


def normalize_header(value: Any) -> str:
    return clean_text(value).strip().lower().replace(" ", "").replace("_", "")


COMPONENT_HEADER_ALIASES = {
    "name": {"\u540d\u79f0", "\u5143\u5668\u4ef6\u540d\u79f0", "name"},
    "category": {"\u5206\u7c7b", "category"},
    "model": {"\u578b\u53f7", "model"},
    "package": {"\u5c01\u88c5", "package", "footprint"},
    "nominal_value": {"\u6807\u79f0\u503c", "\u503c", "value", "comment"},
    "voltage_rating": {"\u8010\u538b", "voltage", "voltage_rating"},
    "current_rating": {"\u8010\u6d41", "\u7535\u6d41", "current", "current_rating"},
    "power_rating": {"\u529f\u7387", "power", "power_rating"},
    "tolerance": {"\u7cbe\u5ea6", "tolerance"},
    "material_type": {"\u6750\u8d28/\u7c7b\u578b", "\u6750\u8d28", "\u7c7b\u578b", "material", "material_type"},
    "manufacturer": {"\u5382\u5546", "manufacturer"},
    "quantity": {"\u6570\u91cf", "qty", "quantity"},
    "box_name": {"\u76d2\u5b50\u540d\u79f0", "\u6536\u7eb3\u76d2", "box", "box_name"},
    "row": {"\u884c", "row"},
    "col": {"\u5217", "col", "column"},
    "description": {"\u63cf\u8ff0", "\u5907\u6ce8", "description"},
    "tags": {"\u6807\u7b7e", "tags"},
}


BOM_HEADER_ALIASES = {
    "comment": {"comment", "\u6ce8\u91ca", "\u503c", "partname", "part name", "comment/value"},
    "designator": {"designator", "\u4f4d\u53f7", "reference", "ref"},
    "footprint": {"footprint", "\u5c01\u88c5", "package"},
    "quantity": {"quantity", "\u6570\u91cf", "qty"},
    "value": {"value", "\u53c2\u6570"},
    "manufacturer_part": {"manufacturerpart", "\u5236\u9020\u5546\u96f6\u4ef6", "mpn", "partnumber", "part number"},
    "manufacturer": {"manufacturer", "\u5236\u9020\u5546"},
    "supplier_part": {"supplierpart", "\u4f9b\u5e94\u5546\u96f6\u4ef6", "lcscpart"},
    "supplier": {"supplier", "\u4f9b\u5e94\u5546"},
}


def resolve_headers(header_row: list[Any], aliases: dict[str, set[str]]) -> dict[str, int]:
    normalized = {index: normalize_header(value) for index, value in enumerate(header_row)}
    mapping: dict[str, int] = {}
    for field, names in aliases.items():
        normalized_aliases = {normalize_header(name) for name in names}
        for index, value in normalized.items():
            if value in normalized_aliases:
                mapping[field] = index
                break
    return mapping


def workbook_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_csv_rows(data: bytes) -> list[list[Any]]:
    text = decode_csv_bytes(data)
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel_tab if "\t" in sample else csv.excel
    return list(csv.reader(io.StringIO(text), dialect))


def parse_xlsx_rows(data: bytes) -> list[tuple[Any, ...]]:
    workbook = load_workbook(filename=io.BytesIO(data), data_only=True)
    sheet = workbook.active
    return list(sheet.iter_rows(values_only=True))


def component_import_rows(file_storage) -> tuple[list[Any], str]:
    filename = file_storage.filename or ""
    ext = Path(filename).suffix.lower()
    data = file_storage.read()
    if ext == ".csv" or ext == ".txt":
        return parse_csv_rows(data), "csv"
    if ext == ".xlsx":
        return parse_xlsx_rows(data), "xlsx"
    raise InventoryError("only csv and xlsx files are supported")


def parse_component_import(file_storage) -> dict[str, Any]:
    rows, file_type = component_import_rows(file_storage)
    if not rows:
        return {"success": 0, "failed": 0, "errors": [], "file_type": file_type}

    header_map = resolve_headers(list(rows[0]), COMPONENT_HEADER_ALIASES)
    if "name" not in header_map:
        raise InventoryError("missing required header: name")

    def cell(row: Any, field: str, default: Any = "") -> Any:
        index = header_map.get(field)
        if index is None:
            return default
        return row[index] if index < len(row) else default

    errors: list[dict[str, Any]] = []
    success_count = 0
    failed_count = 0

    for row_index, row in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in row):
            continue
        try:
            name = clean_text(cell(row, "name"))
            if not name:
                raise InventoryError("missing component name")

            category_name = clean_text(cell(row, "category"))
            category_id = repo.ensure_category_by_name(category_name) if category_name else None

            box = None
            compartment = None
            box_name = clean_text(cell(row, "box_name"))
            if box_name:
                box = repo.find_box_by_name(box_name)
                if not box:
                    raise InventoryError(f"box not found: {box_name}")
                grid_row = clean_optional_int(cell(row, "row"))
                grid_col = clean_optional_int(cell(row, "col"))
                if grid_row is not None and grid_col is not None:
                    compartment = repo.find_compartment_by_coordinates(box["id"], grid_row, grid_col)
                    if not compartment:
                        raise InventoryError(f"compartment not found: {box_name} {grid_row}-{grid_col}")

            payload = {
                "name": name,
                "category_id": category_id,
                "model": clean_text(cell(row, "model")),
                "package": clean_text(cell(row, "package")),
                "nominal_value": clean_text(cell(row, "nominal_value")),
                "voltage_rating": clean_text(cell(row, "voltage_rating")),
                "current_rating": clean_text(cell(row, "current_rating")),
                "power_rating": clean_text(cell(row, "power_rating")),
                "tolerance": clean_text(cell(row, "tolerance")),
                "material_type": clean_text(cell(row, "material_type")),
                "manufacturer": clean_text(cell(row, "manufacturer")),
                "quantity": clean_int(cell(row, "quantity"), 0),
                "description": clean_text(cell(row, "description")),
                "tags": split_tags(cell(row, "tags")) if "tags" in header_map else [],
                "box_id": box["id"] if box else None,
                "compartment_id": compartment["id"] if compartment else None,
                "cell_row": compartment["id"] if compartment else None,
                "cell_col": compartment["col"] if compartment else None,
            }
            repo.create_component(payload, source="import")
            success_count += 1
        except Exception as exc:
            failed_count += 1
            errors.append({"row": row_index, "reason": str(exc)})

    file_logger.write_backend(
        "COMPONENT",
        f"component {file_type} import finished: success={success_count}, failed={failed_count}",
    )
    return {"success": success_count, "failed": failed_count, "errors": errors, "file_type": file_type}


def export_components_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "components"
    sheet.append(
        [
            "\u540d\u79f0",
            "\u5206\u7c7b",
            "\u578b\u53f7",
            "\u5c01\u88c5",
            "\u6807\u79f0\u503c",
            "\u8010\u538b",
            "\u8010\u6d41",
            "\u529f\u7387",
            "\u7cbe\u5ea6",
            "\u6750\u8d28/\u7c7b\u578b",
            "\u5382\u5546",
            "\u6570\u91cf",
            "\u6536\u7eb3\u76d2",
            "\u884c",
            "\u5217",
            "\u63cf\u8ff0",
            "\u6807\u7b7e",
        ]
    )

    page = 1
    while True:
        result = repo.list_components({"page": page, "page_size": 500})
        for item in result["items"]:
            sheet.append(
                [
                    item.get("name"),
                    item.get("category_name"),
                    item.get("model"),
                    item.get("package"),
                    item.get("nominal_value"),
                    item.get("voltage_rating"),
                    item.get("current_rating"),
                    item.get("power_rating"),
                    item.get("tolerance"),
                    item.get("material_type"),
                    item.get("manufacturer"),
                    item.get("quantity"),
                    item.get("box_name"),
                    item.get("grid_row"),
                    item.get("grid_col"),
                    item.get("description"),
                    ", ".join(item.get("tags") or []),
                ]
            )
        if len(result["items"]) < result["page_size"]:
            break
        page += 1
    return workbook_bytes(workbook)


def decode_csv_bytes(data: bytes) -> str:
    detected = chardet.detect(data)
    candidates = [detected.get("encoding"), "utf-8-sig", "utf-8", "utf-16", "gbk"]
    for encoding in candidates:
        if not encoding:
            continue
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def parse_bom_csv(file_storage) -> dict[str, Any]:
    raw = file_storage.read()
    text = decode_csv_bytes(raw)
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel_tab if "\t" in sample else csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    rows = list(reader)
    if not rows:
        return {"project_name": Path(file_storage.filename or "bom").stem, "items": []}

    header_map = resolve_headers(rows[0], BOM_HEADER_ALIASES)
    candidates = repo.bom_component_candidates()
    project_name = Path(file_storage.filename or "bom").stem
    items: list[dict[str, Any]] = []
    matched_count = 0

    def value_score(candidate: dict[str, Any], normalized_query: str, raw_query: str) -> int:
        if not normalized_query and not raw_query:
            return 0
        if normalized_query and candidate["normalized_value"] == normalized_query:
            return 4
        if normalized_query and normalized_query in candidate["search_blob"]:
            return 3
        if normalized_query and normalized_query in candidate["normalized_value"]:
            return 2
        if raw_query and raw_query.lower() in candidate["search_blob"]:
            return 1
        return 0

    for row in rows[1:]:
        if not any(clean_text(value) for value in row):
            continue
        comment = clean_text(row[header_map["comment"]]) if "comment" in header_map else ""
        designator = clean_text(row[header_map["designator"]]) if "designator" in header_map else ""
        footprint = clean_text(row[header_map["footprint"]]) if "footprint" in header_map else ""
        quantity_needed = clean_int(row[header_map["quantity"]], 1) if "quantity" in header_map else 1
        value = clean_text(row[header_map["value"]]) if "value" in header_map else ""

        raw_query = value or comment
        normalized_query = normalize_value(raw_query)
        normalized_package = normalize_footprint(footprint)

        ranked_exact = []
        ranked_fallback = []
        for candidate in candidates:
            score = value_score(candidate, normalized_query, raw_query)
            if score <= 0:
                continue
            if normalized_package and candidate["normalized_package"] == normalized_package:
                ranked_exact.append((score, candidate["quantity"], candidate))
            else:
                ranked_fallback.append((score, candidate["quantity"], candidate))

        ranked_exact.sort(key=lambda item: (item[0], item[1]), reverse=True)
        ranked_fallback.sort(key=lambda item: (item[0], item[1]), reverse=True)

        matched = None
        match_level = None
        if ranked_exact:
            matched = ranked_exact[0][2]
            match_level = "footprint_value"
        elif ranked_fallback:
            matched = ranked_fallback[0][2]
            match_level = "value_only"

        matched_payload = None
        if matched:
            matched_count += 1
            matched_payload = {
                "id": matched["id"],
                "name": matched["name"],
                "model": matched["model"],
                "package": matched["package"],
                "quantity": matched["quantity"],
                "box_id": matched["box_id"],
                "box_name": matched["box_name"],
                "cell_label": matched["cell_label"],
                "row": matched["grid_row"],
                "col": matched["grid_col"],
                "position_x": matched["position_x"],
                "position_y": matched["position_y"],
            }

        items.append(
            {
                "designator": designator,
                "designators": [part.strip() for part in designator.split(",") if part.strip()],
                "comment": comment or value,
                "footprint": footprint,
                "quantity_needed": quantity_needed,
                "matched": matched_payload,
                "match_level": match_level,
            }
        )

    file_logger.write_backend(
        "SYSTEM",
        f"bom import finished: project={project_name}, total={len(items)}, matched={matched_count}, unmatched={len(items) - matched_count}",
    )
    return {
        "project_name": project_name,
        "total_types": len(items),
        "matched": matched_count,
        "unmatched": len(items) - matched_count,
        "items": items,
    }


def export_bom_xlsx(payload: dict[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "picklist"
    sheet.append(
        [
            "BOM\u4f4d\u53f7",
            "\u5143\u5668\u4ef6\u503c",
            "\u5c01\u88c5",
            "\u9700\u8981\u6570\u91cf",
            "\u5e93\u5b58\u6570\u91cf",
            "\u76d2\u5b50",
            "\u683c\u5b50\u4f4d\u7f6e",
            "\u72b6\u6001",
        ]
    )

    for item in payload.get("items") or []:
        matched = item.get("matched") or {}
        quantity_needed = clean_int(item.get("quantity_needed"), 0)
        stock_quantity = clean_int(matched.get("quantity"), 0)
        if not item.get("match_level"):
            status = "unmatched"
        elif stock_quantity < quantity_needed:
            status = "shortage"
        else:
            status = "ok"
        sheet.append(
            [
                item.get("designator"),
                item.get("comment"),
                item.get("footprint"),
                quantity_needed,
                stock_quantity if matched else "",
                matched.get("box_name"),
                matched.get("cell_label"),
                status,
            ]
        )
    return workbook_bytes(workbook)


@app.errorhandler(Exception)
def handle_error(exc: Exception):
    if request.path.startswith("/api"):
        if isinstance(exc, InventoryError):
            return failure(str(exc), 400)
        if isinstance(exc, HTTPException):
            return failure(exc.description, exc.code or 500)
        file_logger.write_backend("SYSTEM", f"exception: {exc}")
        return failure(str(exc), 500)
    if isinstance(exc, HTTPException):
        return Response(exc.description, status=exc.code or 500, mimetype="text/plain")
    file_logger.write_backend("SYSTEM", f"page exception: {exc}")
    return Response("Internal Server Error", status=500, mimetype="text/plain")


@app.get("/")
def index_page():
    return send_from_directory(STATIC_DIR, "index.html")


@app.get("/components/qr-labels")
def component_qr_labels_page():
    ids = parse_component_ids(request.args.get("ids", ""))
    if ids:
        components = [repo.get_component(component_id) for component_id in ids]
    else:
        filters = request.args.to_dict()
        filters["page"] = clean_int(filters.get("page"), 1)
        filters["page_size"] = clean_int(filters.get("page_size"), 500)
        components = repo.list_components(filters)["items"]
    return Response(qr_label_print_page(components), mimetype="text/html")


@app.get("/favicon.ico")
def favicon():
    return Response(status=204)


@app.get("/box/<int:box_id>")
def nfc_box_page(box_id: int):
    return send_from_directory(STATIC_DIR, "box.html")


@app.get("/uploads/<path:filename>")
def uploaded_file(filename: str):
    return send_from_directory(UPLOAD_FOLDER, filename)


@app.get("/api/categories")
def api_categories():
    return success(repo.list_categories())


@app.post("/api/categories")
def api_create_category():
    return success(repo.create_category(request.get_json(silent=True) or {}))


@app.get("/api/cabinets")
def api_cabinets():
    return success(repo.list_cabinets())


@app.post("/api/cabinets")
def api_create_cabinet():
    return success(repo.create_cabinet(request.get_json(silent=True) or {}))


@app.get("/api/cabinets/<int:cabinet_id>")
def api_cabinet_detail(cabinet_id: int):
    return success(repo.get_cabinet(cabinet_id))


@app.put("/api/cabinets/<int:cabinet_id>")
def api_update_cabinet(cabinet_id: int):
    return success(repo.update_cabinet(cabinet_id, request.get_json(silent=True) or {}))


@app.delete("/api/cabinets/<int:cabinet_id>")
def api_delete_cabinet(cabinet_id: int):
    repo.delete_cabinet(cabinet_id)
    return success({"id": cabinet_id})


@app.put("/api/cabinets/<int:cabinet_id>/layout")
def api_update_cabinet_layout(cabinet_id: int):
    return success(repo.update_cabinet_layout(cabinet_id, request.get_json(silent=True) or {}))


@app.get("/api/boxes")
def api_boxes():
    return success(repo.list_boxes())


@app.post("/api/boxes")
def api_create_box():
    return success(repo.create_box(request.get_json(silent=True) or {}))


@app.get("/api/boxes/<int:box_id>")
def api_box_detail(box_id: int):
    return success(repo.get_box(box_id))


@app.put("/api/boxes/<int:box_id>")
def api_update_box(box_id: int):
    return success(repo.update_box(box_id, request.get_json(silent=True) or {}))


@app.delete("/api/boxes/<int:box_id>")
def api_delete_box(box_id: int):
    repo.delete_box(box_id)
    return success({"id": box_id})


@app.put("/api/boxes/<int:box_id>/layout")
def api_update_box_layout(box_id: int):
    return success(repo.update_box_layout(box_id, request.get_json(silent=True) or {}))


@app.get("/api/boxes/<int:box_id>/grid")
def api_box_grid(box_id: int):
    return success(repo.get_box_grid(box_id))


@app.get("/api/components")
def api_components():
    return success(repo.list_components(request.args.to_dict()))


@app.post("/api/components")
def api_create_component():
    return success(repo.create_component(request.get_json(silent=True) or {}))


@app.get("/api/components/<int:component_id>")
def api_component_detail(component_id: int):
    return success(repo.get_component(component_id))


@app.put("/api/components/<int:component_id>")
def api_update_component(component_id: int):
    return success(repo.update_component(component_id, request.get_json(silent=True) or {}))


@app.delete("/api/components/<int:component_id>")
def api_delete_component(component_id: int):
    repo.delete_component(component_id)
    return success({"id": component_id})


@app.get("/api/components/<int:component_id>/qr.svg")
def api_component_qr_svg(component_id: int):
    repo.get_component(component_id)
    return Response(
        component_qr_svg(component_id),
        mimetype="image/svg+xml",
        headers={"Cache-Control": "no-store"},
    )


@app.post("/api/components/import")
def api_import_components():
    file = request.files.get("file")
    if not file:
        raise InventoryError("missing upload file")
    result = parse_component_import(file)
    return success(result)


@app.get("/api/components/export")
def api_export_components():
    data = export_components_xlsx()
    file_logger.write_backend("COMPONENT", "export components excel")
    return excel_response(data, "components_export.xlsx")


@app.post("/api/images/upload")
def api_upload_image():
    file = request.files.get("file")
    if not file or not file.filename:
        raise InventoryError("missing image file")
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_IMAGE_EXTENSIONS:
        raise InventoryError("only jpg/png/gif/webp images are allowed")

    component_id = clean_optional_int(request.form.get("component_id"))
    is_primary = str(request.form.get("is_primary") or "").lower() in {"1", "true", "yes"}
    filename = f"{uuid.uuid4().hex}{ext}"
    thumb_filename = f"thumbnails/{uuid.uuid4().hex}{ext}"

    target = UPLOAD_FOLDER / filename
    thumb_target = UPLOAD_FOLDER / thumb_filename
    target.parent.mkdir(parents=True, exist_ok=True)
    thumb_target.parent.mkdir(parents=True, exist_ok=True)
    file.save(target)

    with Image.open(target) as image:
        image = ImageOps.exif_transpose(image)
        image.thumbnail((200, 200))
        image.save(thumb_target)

    return success(repo.create_image_record(filename, thumb_filename, component_id, is_primary))


@app.delete("/api/images/<int:image_id>")
def api_delete_image(image_id: int):
    repo.delete_image(image_id)
    return success({"id": image_id})


@app.post("/api/documents/upload")
def api_upload_document():
    file = request.files.get("file")
    if not file or not file.filename:
        raise InventoryError("missing document file")
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_DOCUMENT_EXTENSIONS:
        raise InventoryError("only pdf/doc/xls/ppt/txt/csv/zip documents are allowed")

    component_id = clean_optional_int(request.form.get("component_id"))
    original_name = clean_text(Path(file.filename).name) or secure_filename(file.filename) or f"document{ext}"
    stored_name = f"{uuid.uuid4().hex}{ext}"
    filename = f"documents/{stored_name}"
    target = DOCUMENT_FOLDER / stored_name
    target.parent.mkdir(parents=True, exist_ok=True)
    file.save(target)

    return success(
        repo.create_document_record(
            filename,
            original_name,
            file.mimetype,
            target.stat().st_size,
            component_id,
        )
    )


@app.delete("/api/documents/<int:document_id>")
def api_delete_document(document_id: int):
    repo.delete_document(document_id)
    return success({"id": document_id})


@app.post("/api/frontend/log")
def api_frontend_log():
    payload = request.get_json(silent=True) or {}
    logs = payload.get("logs")
    if not isinstance(logs, list):
        raise InventoryError("logs must be a list")
    if len(logs) > 100:
        raise InventoryError("at most 100 logs per request")
    written = file_logger.write_frontend_batch(logs)
    return success({"written": written})


@app.get("/api/settings/tokens")
def api_settings_tokens():
    return success(repo.list_api_tokens())


@app.post("/api/settings/tokens")
def api_settings_create_token():
    return success(repo.create_api_token(request.get_json(silent=True) or {}))


@app.delete("/api/settings/tokens/<int:token_id>")
def api_settings_delete_token(token_id: int):
    repo.delete_api_token(token_id)
    return success({"id": token_id})


@app.post("/api/stock/in")
def api_stock_in():
    return success(repo.record_stock("in", request.get_json(silent=True) or {}))


@app.post("/api/stock/out")
def api_stock_out():
    return success(repo.record_stock("out", request.get_json(silent=True) or {}))


@app.get("/api/stock/logs/<int:component_id>")
def api_stock_logs(component_id: int):
    page = clean_int(request.args.get("page"), 1)
    page_size = clean_int(request.args.get("page_size"), 20)
    return success(repo.list_stock_logs(component_id, page, page_size))


@app.get("/api/stats")
def api_stats():
    return success(repo.get_stats())


@app.get("/api/map")
def api_map():
    return success(repo.get_map_data())


@app.post("/api/map/search")
def api_map_search():
    payload = request.get_json(silent=True) or {}
    keyword = clean_text(payload.get("keyword"))
    return success({"results": repo.search_map(keyword)})


@app.post("/api/nfc/bind")
def api_nfc_bind():
    payload = request.get_json(silent=True) or {}
    box_id = clean_int(payload.get("box_id"))
    uid = clean_text(payload.get("uid"))
    return success(repo.bind_nfc(box_id, uid))


@app.post("/api/nfc/write")
def api_nfc_write():
    payload = request.get_json(silent=True) or {}
    box_id = clean_int(payload.get("box_id"))
    return success(repo.get_nfc_payload(box_id, SERVER_URL))


@app.get("/api/nfc/lookup/<string:uid>")
def api_nfc_lookup(uid: str):
    return success(repo.lookup_nfc(uid))


@app.get("/api/box/<int:box_id>")
def api_nfc_box_data(box_id: int):
    return success(repo.get_box_grid(box_id))


@app.post("/api/bom/import")
def api_bom_import():
    file = request.files.get("file")
    if not file:
        raise InventoryError("missing csv file")
    return success(parse_bom_csv(file))


@app.post("/api/bom/consume")
def api_bom_consume():
    payload = request.get_json(silent=True) or {}
    return success(repo.consume_bom(payload))


@app.post("/api/bom/export")
def api_bom_export():
    payload = request.get_json(silent=True) or {}
    data = export_bom_xlsx(payload)
    file_logger.write_backend("SYSTEM", "export bom picklist")
    return excel_response(data, "bom_picklist.xlsx")


if __name__ == "__main__":
    file_logger.write_backend("SYSTEM", f"service start: http://{HOST}:{PORT}")
    app.run(host=HOST, port=PORT, debug=True)
