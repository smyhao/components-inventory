from __future__ import annotations

import csv
import io
from typing import Any

import chardet
from openpyxl import Workbook

from models import clean_text


def normalize_header(value: Any) -> str:
    return clean_text(value).strip().lower().replace(" ", "").replace("_", "")


def resolve_headers(header_row: list[Any], aliases: dict[str, set[str]]) -> dict[str, int]:
    normalized = {index: normalize_header(value) for index, value in enumerate(header_row)}
    mapping: dict[str, int] = {}
    for field, names in aliases.items():
        normalized_aliases = {normalize_header(name) for name in names}
        for index, value in normalized.items():
            if value in normalized_aliases:
                mapping[field] = index
                break
    return mapping


def workbook_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def decode_csv_bytes(data: bytes) -> str:
    detected = chardet.detect(data)
    candidates = [detected.get("encoding"), "utf-8-sig", "utf-8", "utf-16", "gbk"]
    for encoding in candidates:
        if not encoding:
            continue
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def parse_csv_rows(data: bytes) -> list[list[Any]]:
    text = decode_csv_bytes(data)
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel_tab if "\t" in sample else csv.excel
    return list(csv.reader(io.StringIO(text), dialect))


def parse_xlsx_rows(data: bytes) -> list[tuple[Any, ...]]:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=io.BytesIO(data), data_only=True)
    sheet = workbook.active
    return list(sheet.iter_rows(values_only=True))
